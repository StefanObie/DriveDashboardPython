import pandas as pd
import urllib.parse
import requests
import config
import openpyxl
import glob
import os

MILES_TO_KM = 1.60934

def ceil_time_to_minute(time):
    return time + pd.Timedelta(minutes=1) - pd.Timedelta(seconds=time.second, microseconds=time.microsecond)

def load_file():
    # Read the latest XLS file from the MovementReports directory
    movement_reports_dir = 'DetailedReports'
    xls_files = glob.glob(os.path.join(movement_reports_dir, '*.xls*'))
    if not xls_files:
        raise FileNotFoundError(f"No XLS files found in {movement_reports_dir}")
    file_path = max(xls_files, key=lambda f: os.path.getctime(os.path.abspath(f)))

    # Specific file paths for testing
    # file_path = 'MovementReports\\StefanMovementReport17Apr25.csv'
    # file_path = 'MovementReports\\StefanMovementReportApr2025.csv' # 1-3 April missing (max data retention 38 days)
    # file_path = 'MovementReports\\StefanMovementReport29May2025.csv' # 29-31 May missing
    # file_path = 'MovementReports\\StefanMovementReport30May2025.csv' # 31 May missing
    # file_path = 'MovementReports\\StefanMovementReportMay2025.csv'

    print(f"Loading file: {file_path}\n")
    df = pd.read_excel(file_path)
    # Handle merged cells by forward-filling NaN values
    # df = df.fillna(method='ffill')
    return df

def preprocessing(df):
    # Remove the report header (first 17 rows)
    df = df.iloc[17:].reset_index(drop=True)

    # Keep columns - 2 (Date), 8 (Event), 11 (Location), 20 (Speed), 21 (Distance)
    df = df[[df.columns[2], df.columns[8], df.columns[11], df.columns[20], df.columns[21]]]
    df.columns = ['Date', 'Event', 'Location', 'Speed', 'Distance']

    # Split the 'Location' column into 'Longitude' and 'Latitude'
    df[['Longitude', 'Latitude']] = df['Location'].str.extract(r'Long\s*:\s*([\d\,\-]+)\,\s*Lat\s*:\s*([\d\,\-]+)')
    df[['Longitude', 'Latitude']] = df[['Longitude', 'Latitude']].replace(',', '.', regex=True).astype(float)
    df = df.drop(columns=['Location'], errors='ignore')

    # Convert Date column to datetime objects (Format: 2025/08/25 18:24)
    df['DateTime'] = pd.to_datetime(df['Date'], format='%Y/%m/%d %H:%M', errors='coerce')

    # Trip Numbering
    df['TripNumber'] = (df['Event'] == 'Start up').cumsum()

    return df

def report_dates(df, full_month=False):
    df_report_date = df.copy()
    df_report_date['Date'] = df_report_date['DateTime'].dt.date

    if (full_month): # Full Month
        first_date = df_report_date['Date'].min().replace(day=1)
        days_in_month = pd.Period(first_date, freq='M').days_in_month
        last_date = first_date + pd.Timedelta(days=days_in_month -1)
    else: # Month-to-Date (Last day as in the Report)
        # first_date = df_no_drive['Date'].min() # Report Start Date
        first_date = df_report_date['Date'].min().replace(day=1) # Month-to-Date
        last_date = df_report_date['Date'].max()
    print(f"Report Dates: {first_date} to {last_date}")

    return first_date, last_date

def no_drive_days(df, first_date, last_date):
    df_no_drive = df.copy()
    df_no_drive['Date'] = df_no_drive['DateTime'].dt.date
    df_no_drive = df_no_drive[df["Event"] == 'Start up']
    df_no_drive = df_no_drive.drop_duplicates(subset=['Date'], keep='first')

    num_no_drive_days = (last_date - first_date).days + 1 - len(df_no_drive)
    print(f"No-Drive Days: {num_no_drive_days}\n")
    return num_no_drive_days

def driving_violations(df, violation='Harsh Braking'):
    df_braking = df[df['Event'] == violation]
    df_braking = df_braking.sort_values(by='DateTime') 
    df_braking = df_braking[~(df_braking['DateTime'].diff().dt.total_seconds().abs() <= 120)] # Remove false alarms for Harsh Braking
    print(f"{violation}: {len(df_braking)} ({len(df_braking) *8} Points)\n")

    return len(df_braking) *8

def get_speed_limit(lat, lon):
    # Get the speed limit for a given latitude and longitude using HERE API.
    base_url = "https://revgeocode.search.hereapi.com/v1/revgeocode"
    params = {
        "at": f"{lat},{lon},50",
        "maxResults": "1",
        "apiKey": config.HERE_API_KEY,
        "showNavAttributes": "speedLimits",
        "types": "street"
    }
    
    url = f"{base_url}?{urllib.parse.urlencode(params)}"
    print(f"\tCalling HERE API...")
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        if 'items' in data and len(data['items']) > 0:
            speed_limits = data['items'][0].get('navigationAttributes', {}).get('speedLimits', [])
            if speed_limits:
                # Get all maxSpeed values and return the smallest one
                max_speeds = [sl.get('maxSpeed') for sl in speed_limits if 'maxSpeed' in sl]
                if max_speeds:
                    return min(max_speeds)
    return None

def speed_violation(df, call_here_api_for_speedlimit=False, default_speed_limit=60):
    df_speed = df[(df['Event'] == 'Speed Violation') & 
                (df['Speed'] >= default_speed_limit+10)] # Everything under 70 km/h is not a violation
    
    print(f"Speed Violations") # Heading
    
    if len(df_speed) > 10 and call_here_api_for_speedlimit: # Do not call HERE API if there are too many speed violations
        print(df_speed[['DateTime', 'Speed', 'Latitude', 'Longitude']])
        confirm = input(f"There are {len(df_speed)} speed violation(s). Type y (yes) to continue with HERE API calls: ")
        if confirm.strip().lower() != 'y':
            print(f"Using default speed limit of {default_speed_limit} km/h.")
            call_here_api_for_speedlimit = False

    speed_penalty_total = 0
    for _, row in df_speed.iterrows():
        print(f"{row['DateTime'].strftime('%Y/%m/%d %H:%M')} @ {row['Latitude']}, {row['Longitude']}")

        speed = row['Speed']
        penalty = 0

        if call_here_api_for_speedlimit:
            speedlimit = get_speed_limit(row['Latitude'], row['Longitude'])
        else: # Use a fixed speed limit
            speedlimit = default_speed_limit # Default speed limit

        if speedlimit is None:
            print(f"[ERROR] No speed limit found for coordinates: {row['Latitude']}, {row['Longitude']}")
            continue # Skip to the next iteration

        if speed - speedlimit >= 10:
            if speed - speedlimit <= 15: # 10..15
                penalty = 3
            elif speed - speedlimit <= 25: # 16..25
                penalty = 8
            elif speed - speedlimit > 25: # 26..
                penalty = 15

            speed_penalty_total += penalty
            
        print(f"\tSpeed Violation: {speed} km/h, Speed Limit: {speedlimit} km/h, Penalty: {penalty} points")

    print(f"Speed Violation Total: {speed_penalty_total} Points\n")
    return speed_penalty_total

def night_time_driving(df):
    # Night Time Driving
    df['Hour'] = df['DateTime'].dt.hour
    # Only include records between 23:00 and 04:30 (not greater than 04:30)
    df_night_driving = df[
        ((df['Hour'] >= 23) | 
         ((df['Hour'] < 4) | ((df['Hour'] == 4) & (df['DateTime'].dt.minute <= 30)))) &
        (df['Event'] != 'Health Check; (Ignition off)')
    ].copy()
        
    print(f"Night Time Driving")
    night_penalty_total = 0

    # For each TripNumber in the night driving dataframe
    trip_nums = df_night_driving['TripNumber'].unique()
    for trip in trip_nums:
        trip_df = df[df['TripNumber'] == trip]

        # Find difference between "Start up" and "Ignition off"
        if len(trip_df) > 1:
            start_time = trip_df[trip_df['Event'] == 'Start up']['DateTime'].min()
            end_time = trip_df[trip_df['Event'] == 'Ignition off']['DateTime'].max()
            duration = end_time - start_time

            # If time is between 23h and 4h30, add a penalty for each minute.
            t = start_time + pd.Timedelta(minutes=1, seconds=-1)
            night_penalty = 0
            while t <= end_time:
                if t.hour in [23, 4]:
                    night_penalty += 2
                elif t.hour in [0, 3]:
                    night_penalty += 4
                elif t.hour in [1, 2]:
                    night_penalty += 6
                # print(t, night_penalty)
                t += pd.Timedelta(minutes=1)
            # print(night_penalty)

            night_penalty_total += night_penalty 
            print(f"\tTrip Number: {trip}, Duration: {duration.components.hours}h:{duration.components.minutes:02d}m, Night Penalty: {night_penalty} points")
            # print(trip_df[['DateTime', 'VehicleStatus']])
            
    print(f"Night Time Driving Total: {night_penalty_total}\n")
    return night_penalty_total

def distance(df):
    d = df.iloc[7, 4] * MILES_TO_KM
    print(f"Month-to-Date Distance: {d:.0f} km")
    return d

def write_to_excel(last_date, drive_pen=0, night_pen=0, no_drive=0, dist=0, sheetname='DEV'):
    output_file = 'DriveTemplateDevelopmentPython.xlsm' if sheetname == 'DEV' else 'DriveSummaryPython.xlsx'

    try:
        wb = openpyxl.load_workbook(output_file)
    except FileNotFoundError:
        print(f"File {output_file} not found.")
        return

    if sheetname in wb.sheetnames:
        ws = wb[sheetname]
    else:
        print(f"Sheet {sheetname} not found in the workbook.")
        return

    ws['J6'] = drive_pen
    ws['J7'] = night_pen
    ws['J8'] = no_drive
    ws['J9'] = dist
    ws['J14'] = last_date

    wb.save(output_file)
    wb.close()

    print(f"Data written to Excel successfully.")

def main():
    # Config
    full_month = False
    call_here_api_for_speedlimit = True
    dev = False
    save_to_excel = True

    # Load and preprocess data
    df = load_file()
    dist = distance(df)
    df = preprocessing(df)

    first_date, last_date = report_dates(df, full_month=full_month)
    no_drive = no_drive_days(df, first_date, last_date)

    drive_pen = 0
    drive_pen += driving_violations(df, violation='Harsh Braking')
    drive_pen += driving_violations(df, violation='Harsh Acceleration')
    drive_pen += driving_violations(df, violation='Harsh Cornering')
    drive_pen += speed_violation(df, call_here_api_for_speedlimit)
    
    night_pen = night_time_driving(df)

    sheetname = 'DEV' if dev else 'Stefan'
    if save_to_excel:
        write_to_excel(last_date, drive_pen, night_pen, no_drive, dist, sheetname)

if __name__ == '__main__':
    main()